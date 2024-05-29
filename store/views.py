from django.http import HttpResponse,JsonResponse
from django.shortcuts import get_object_or_404, render, redirect
from django.urls import reverse
from django.views import View
from django.views.generic import ListView
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.db.models import Sum

from .models import Yetkazib_beruvchilar, Chiqimlar, Kategoriyalar, Mahsulotlar, Kirimlar,Omborxona
from .forms import YetkazibBeruvchilarForm,ChiqimForm,MahsulotlarForm,KirimForm,KategoriyalarForm

import openpyxl
from openpyxl.styles import Font

# Yetkazib beruvchi views
@login_required(login_url='login')
def create_yetkazib_beruvchi(request):
    form = YetkazibBeruvchilarForm()
    if request.method == 'POST':
        form = YetkazibBeruvchilarForm(request.POST)
        if form.is_valid():
            yetkazib_beruvchi = form.save(commit=False)
            yetkazib_beruvchi.user = request.user
            yetkazib_beruvchi.save()
            return redirect('yetkazib-beruvchi-list')
    context = {
        'form': form
    }
    return render(request, 'store/create_yetkazib_beruvchi.html', context)

class YetkazibBeruvchiListView(ListView):
    model = Yetkazib_beruvchilar
    template_name = 'store/yetkazib_beruvchi_list.html'
    context_object_name = 'yetkazib_beruvchilar'

class YetkazibBeruvchilarXLSXDownloadView(View):
    def get(self, request, *args, **kwargs):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Поставшики'

        headers = ['#', 'Ф.И.О', 'ИНН', 'Адрес', 'Дата']
        sheet.append(headers)

        for cell in sheet["1:1"]:
            cell.font = Font(bold=True)

        yetkazib_beruvchilar = Yetkazib_beruvchilar.objects.all()

        for idx, yetkazib_beruvchi in enumerate(yetkazib_beruvchilar, start=1):
            sheet.append([
                idx,
                yetkazib_beruvchi.FISH,
                yetkazib_beruvchi.INN,
                yetkazib_beruvchi.address,
                yetkazib_beruvchi.created_date.strftime('%Y-%m-%d')  # Format the date
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=Postavshiki.xlsx'

        workbook.save(response)
        return response
    
class YetkazibBeruvchilarEditView(View):
    def get(self, request, id):
        yetkazib_beruvchi = get_object_or_404(Yetkazib_beruvchilar, id=id)
        form = YetkazibBeruvchilarForm(instance=yetkazib_beruvchi)
        return render(request, 'edit/edit_yetkazib_beruvchi.html', {'form': form})

    def post(self, request, id):
        yetkazib_beruvchi = get_object_or_404(Yetkazib_beruvchilar, id=id)
        form = YetkazibBeruvchilarForm(request.POST, instance=yetkazib_beruvchi)
        if form.is_valid():
            form.save()
            return redirect('yetkazib-beruvchi-list')
        return render(request, 'edit/edit_yetkazib_beruvchi.html', {'form': form})

class YetkazibBeruvchilarDeleteView(View):
    def get(self, request, id):
        yetkazib_beruvchi = get_object_or_404(Yetkazib_beruvchilar, id=id)
        yetkazib_beruvchi.delete()
        return redirect('yetkazib-beruvchi-list')
















# Chiqim views
@login_required(login_url='login')
def create_chiqim(request):
    form = ChiqimForm()
    if request.method == 'POST':
        form = ChiqimForm(request.POST)
        if form.is_valid():
            chiqim = form.save(commit=False)
            chiqim.user = request.user

            mah_v= str(form.cleaned_data['mahsulot'])
            mah_chiqim = mah_v.split(' ', 1)[0]
            mah_ombor = Omborxona.objects.filter(mahsulot__nomi=mah_chiqim).first()
            miqdor = form.cleaned_data['miqdori']
            if miqdor>mah_ombor.umumiy_soni:
                return JsonResponse({'error': 'Not enough stock', 'available_stock': mah_ombor.umumiy_soni,'birlik':mah_ombor.mahsulot.ulchov_birligi})


            chiqim.save()

            mah_v= str(form.cleaned_data['mahsulot'])
            mah_chiqim = mah_v.split(' ', 1)[0]
            chiqim_obj = Chiqimlar.objects.filter(mahsulot__nomi=mah_chiqim).first()
            mahsulot = Mahsulotlar.objects.filter(nomi=mah_chiqim).first()

            if chiqim_obj and mahsulot:  
                mah_ombor = Omborxona.objects.filter(mahsulot__nomi=mah_chiqim).first()
                if mah_ombor:
                    mah_ombor.umumiy_soni -= chiqim.miqdori 
                    narx = chiqim.miqdori * mahsulot.narxi 
                    mah_ombor.umumiy_narx -= narx
                    mah_ombor.save()
            
            return redirect('chiqim-list')
    
    context = {
        'form': form
    }
    return render(request, 'store/create_chiqim.html', context)


class ChiqimListView(ListView):
   def get(self,request):
       chiqim = Chiqimlar.objects.all().order_by('-id')
       return render(request,'store/chiqim_list.html',{'chiqimlar':chiqim})

class ChiqimlarXLSXDownloadView(View):
    def get(self, request, *args, **kwargs):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Расходы'

        headers = ['#', 'Название продукта', 'Количество', 'Имя', 'Фамилия', 'Номер телефона', 'Дата продажи']
        sheet.append(headers)

        for cell in sheet["1:1"]:
            cell.font = Font(bold=True)

        chiqimlar = Chiqimlar.objects.all()

        for idx, chiqim in enumerate(chiqimlar, start=1):
            sheet.append([
                idx,
                str(chiqim.mahsulot),
                chiqim.miqdori,
                chiqim.ism,
                chiqim.familiya,
                chiqim.telefon,
                chiqim.sotilgan_sana.strftime('%Y-%m-%d')  # Format the date
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=Rasxod.xlsx'

        workbook.save(response)
        return response











# Mahsulotlar views
@login_required(login_url='login')
def create_mahsulot(request):
    form = MahsulotlarForm()
    if request.method == 'POST':
        form = MahsulotlarForm(request.POST)
        if form.is_valid():
            mahsulot = form.save(commit=False)
            mahsulot.user = request.user 
            mahsulot.save()
            
            return redirect('mahsulot-list')

    context = {
        'form': form,
    }
    return render(request, 'store/create_mahsulot.html', context)

class MahsulotlarListView(ListView):
    model = Mahsulotlar
    template_name = 'store/mahsulot_list.html'
    context_object_name = 'mahsulotlar'

class MahsulotlarXLSXDownloadView(View):
    def get(self, request, *args, **kwargs):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Продукты'

        headers = ['#', 'Название', 'Категория', 'Цена', 'Единица количества', 'Дата']
        sheet.append(headers)

        for cell in sheet["1:1"]:
            cell.font = Font(bold=True)

        mahsulotlar = Mahsulotlar.objects.all()

        for idx, mahsulot in enumerate(mahsulotlar, start=1):
            sheet.append([
                idx,
                mahsulot.nomi,
                str(mahsulot.kategoriya),
                mahsulot.narxi,
                mahsulot.ulchov_birligi,
                mahsulot.created_date.strftime('%Y-%m-%d')  # Format the date
            ])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename=Products.xlsx'

        workbook.save(response)
        return response

class MahsulotlarEditView(View):
    def get(self, request, id):
        mahsulot = get_object_or_404(Mahsulotlar, id=id)
        form = MahsulotlarForm(instance=mahsulot)
        return render(request, 'edit/edit_mahsulot.html', {'form': form})

    def post(self, request, id):
        mahsulot = get_object_or_404(Mahsulotlar, id=id)
        form = MahsulotlarForm(request.POST, instance=mahsulot)
        if form.is_valid():
            form.save()
            return redirect(reverse("mahsulot-list"))
        return render(request, 'edit/edit_mahsulot.html', {'form': form})

class MahsulotlarDeleteView(View):
    def get(self, request, id):
        mahsulot = get_object_or_404(Mahsulotlar, id=id)
        mahsulot.delete()
        return redirect('mahsulot-list')  

















# Kirimlar views
@login_required(login_url='login')
def create_kirim(request):
    form = KirimForm()
    if request.method == 'POST':
        form = KirimForm(request.POST)
        if form.is_valid():
            kirim = form.save(commit=False)
            kirim.user = request.user
            kirim.save()
            mah_v= str(form.cleaned_data['mahsulot'])
            mah_ombor = mah_v.split(' ', 1)[0]
            kirim_mah=Kirimlar.objects.filter(mahsulot__nomi=mah_ombor)
            mahsulot=Omborxona.objects.filter(mahsulot__nomi=mah_ombor)
            b=[i for i in kirim_mah][-1]
            mak=Mahsulotlar.objects.get(nomi=mah_ombor)
            if mahsulot:
                mahsulot[0].umumiy_soni+=b.soni
                mahsulot[0].umumiy_narx+=(b.narx)*b.soni
                mahsulot[0].save()
            else:
                Omborxona.objects.create(
                    mahsulot=mak,
                    umumiy_soni=b.soni,
                    umumiy_narx=(b.narx)*b.soni+(b.narx)*b.soni
                )
            return redirect('kirim-list')
    context = {
        'form': form
    }
    return render(request, 'store/create_kirim.html', context)

class KirimlarListView(ListView):
    def get(self,request):
        kirim = Kirimlar.objects.all().order_by('-id')
        return render(request, 'store/kirim_list.html', {"kirimlar":kirim})


class KirimlarXLSXDownloadView(View):
    def get(self, request):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Приходы"

        headers = ['#', 'Поставшик', 'Продукт', 'Количество', 'Единица измерения', 'Дата получения продукта']
        worksheet.append(headers)

        for cell in worksheet["1:1"]:
            cell.font = Font(bold=True)

        kirimlar = Kirimlar.objects.all().order_by('-id')
        for idx, kirim in enumerate(kirimlar, start=1):
            worksheet.append([
                idx,
                str(kirim.yetkazib_beruvchi),
                kirim.mahsulot.nomi,
                kirim.soni,
                kirim.mahsulot.ulchov_birligi,
                kirim.keltirilgan_sana,
            ])

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="Prixod.xlsx"'

        workbook.save(response)

        return response